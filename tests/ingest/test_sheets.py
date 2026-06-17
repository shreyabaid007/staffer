"""End-to-end tests for ``ingest_candidates`` (T-004).

Covers I-LOAD-1/2, I-VAL-1, I-SUM-1, I-EDGE-1/2/3, I-DET-1. Synthetic in-memory
workbooks (openpyxl ``Workbook()`` saved under ``tmp_path``) exercise the edge
cases; one smoke test runs the real shipped workbook. No network, no LLM.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from dsm.ingest.models import IngestError
from dsm.ingest.sheets import ingest_candidates
from dsm.models import CandidateSource, FreeNow, NewJoiner, RollingOff

# Canonical headers matching the real workbook (header row = row 2).
BEACH_HEADERS = [
    "#",
    "Name",
    "Email",
    "Grade",
    "Key Skills",
    "Location",
    "Chennai-open",
    "Days on Beach",
    "Notes",
]
ROLLING_HEADERS = [
    "#",
    "Name",
    "Email",
    "Grade",
    "Key Skills",
    "Current Client",
    "Roll-off Date",
    "Confidence",
    "Location",
    "Chennai-open",
    "Notes",
]
JOINER_HEADERS = [
    "#",
    "Name",
    "Email",
    "Grade",
    "Key Skills (from CV)",
    "Join Date",
    "Location",
    "Chennai-open",
    "Notes",
]


def _write_workbook(
    tmp_path: Path, sheets: dict[str, tuple[list[str], list[list[object]]]]
) -> str:
    """Build a workbook: each sheet gets a title row (1), header row (2), then data."""
    wb = Workbook()
    # Drop the default sheet so only the requested tabs exist.
    default = wb.active
    assert default is not None
    wb.remove(default)
    for name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title=name)
        ws.append([f"{name} - synthetic title"])  # row 1
        ws.append(headers)  # row 2
        for row in rows:
            ws.append(row)
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    return str(path)


def test_each_tab_maps_to_its_availability(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path,
        {
            "Beach": (
                BEACH_HEADERS,
                [[1, "A", "a@x.example", "Lead", "Java", "Bengaluru", "Yes", 5, "n"]],
            ),
            "Rolling Off": (
                ROLLING_HEADERS,
                [
                    [
                        1,
                        "B",
                        "b@x.example",
                        "Lead",
                        "Java",
                        "C",
                        "2026-07-01",
                        "high",
                        "Chennai",
                        "No",
                        "n",
                    ]
                ],
            ),
            "New Joiners": (
                JOINER_HEADERS,
                [
                    [
                        1,
                        "C",
                        "c@x.example",
                        "Senior",
                        "Python",
                        "2026-07-11",
                        "Remote (India)",
                        "No",
                        "n",
                    ]
                ],
            ),
        },
    )
    candidates, summary = ingest_candidates(path)

    assert summary.candidate_rows_seen == 3
    assert summary.candidates_ingested == 3
    assert summary.issues == []
    assert isinstance(candidates["a@x.example"].availability, FreeNow)
    assert candidates["a@x.example"].source is CandidateSource.BEACH
    rolling = candidates["b@x.example"]
    assert isinstance(rolling.availability, RollingOff)
    assert rolling.availability.confidence == "high"
    assert isinstance(candidates["c@x.example"].availability, NewJoiner)
    assert candidates["c@x.example"].source is CandidateSource.NEW_JOINER


def test_duplicate_email_first_occurrence_wins(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path,
        {
            "Beach": (
                BEACH_HEADERS,
                [[1, "Beach Bob", "dup@x.example", "Lead", "Java", "Bengaluru", "Yes", 5, "n"]],
            ),
            "Rolling Off": (ROLLING_HEADERS, []),
            "New Joiners": (
                JOINER_HEADERS,
                [
                    [
                        1,
                        "Joiner Bob",
                        "dup@x.example",
                        "Senior",
                        "Python",
                        "2026-07-11",
                        "Bengaluru",
                        "No",
                        "n",
                    ]
                ],
            ),
        },
    )
    candidates, summary = ingest_candidates(path)

    assert summary.candidates_ingested == 1
    assert summary.duplicate_emails_skipped == 1
    assert candidates["dup@x.example"].name == "Beach Bob"  # Beach kept (first in tab order)
    dup_issues = [i for i in summary.issues if "duplicate" in i.reason]
    assert len(dup_issues) == 1
    assert dup_issues[0].sheet == "New Joiners"
    assert "Beach" in dup_issues[0].reason


def test_blank_row_skipped_not_an_issue(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path,
        {
            "Beach": (
                BEACH_HEADERS,
                [
                    [1, "A", "a@x.example", "Lead", "Java", "Bengaluru", "Yes", 5, "n"],
                    # Blank row in the *middle* — openpyxl trims fully-empty trailing rows,
                    # so a real data row must follow it to exercise the blank-skip path.
                    [None, None, None, None, None, None, None, None, None],
                    [2, "B", "b@x.example", "Lead", "Kotlin", "Chennai", "No", 3, "n"],
                ],
            ),
            "Rolling Off": (ROLLING_HEADERS, []),
            "New Joiners": (JOINER_HEADERS, []),
        },
    )
    _candidates, summary = ingest_candidates(path)

    assert summary.candidates_ingested == 2
    assert summary.candidate_rows_seen == 2  # blank row not counted as a candidate row
    assert summary.blank_rows_skipped == 1
    assert summary.issues == []


def test_unparseable_date_dropped_others_ingested(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path,
        {
            "Beach": (BEACH_HEADERS, []),
            "Rolling Off": (ROLLING_HEADERS, []),
            "New Joiners": (
                JOINER_HEADERS,
                [
                    [
                        1,
                        "Bad",
                        "bad@x.example",
                        "Senior",
                        "Python",
                        "7th July",
                        "Bengaluru",
                        "No",
                        "n",
                    ],
                    [
                        2,
                        "Good",
                        "good@x.example",
                        "Senior",
                        "Python",
                        "2026-07-11",
                        "Bengaluru",
                        "No",
                        "n",
                    ],
                ],
            ),
        },
    )
    candidates, summary = ingest_candidates(path)

    assert "good@x.example" in candidates
    assert "bad@x.example" not in candidates
    assert summary.candidate_rows_seen == 2
    assert summary.candidates_ingested == 1
    assert len(summary.issues) == 1
    assert summary.issues[0].email == "bad@x.example"
    assert "date" in summary.issues[0].reason


def test_bad_confidence_is_an_issue(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path,
        {
            "Beach": (BEACH_HEADERS, []),
            "Rolling Off": (
                ROLLING_HEADERS,
                [
                    [
                        1,
                        "X",
                        "x@x.example",
                        "Lead",
                        "Java",
                        "C",
                        "2026-07-01",
                        "maybe",
                        "Chennai",
                        "No",
                        "n",
                    ]
                ],
            ),
            "New Joiners": (JOINER_HEADERS, []),
        },
    )
    candidates, summary = ingest_candidates(path)

    assert candidates == {}
    assert len(summary.issues) == 1
    assert "confidence" in summary.issues[0].reason


def test_missing_required_header_raises(tmp_path: Path) -> None:
    headers_without_email = [
        "#",
        "Name",
        "Grade",
        "Key Skills",
        "Location",
        "Chennai-open",
        "Days on Beach",
        "Notes",
    ]
    path = _write_workbook(
        tmp_path,
        {
            "Beach": (headers_without_email, []),
            "Rolling Off": (ROLLING_HEADERS, []),
            "New Joiners": (JOINER_HEADERS, []),
        },
    )
    with pytest.raises(IngestError, match="Email"):
        ingest_candidates(path)


def test_missing_supply_tab_raises(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path,
        {
            "Beach": (BEACH_HEADERS, []),
            # "Rolling Off" omitted
            "New Joiners": (JOINER_HEADERS, []),
        },
    )
    with pytest.raises(IngestError, match="Rolling Off"):
        ingest_candidates(path)


def test_non_supply_tab_is_ignored(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path,
        {
            "Open Roles": (["Role ID", "Title"], [["ROLE-01", "Engineer"]]),
            "Beach": (
                BEACH_HEADERS,
                [[1, "A", "a@x.example", "Lead", "Java", "Bengaluru", "Yes", 5, "n"]],
            ),
            "Rolling Off": (ROLLING_HEADERS, []),
            "New Joiners": (JOINER_HEADERS, []),
        },
    )
    candidates, summary = ingest_candidates(path)
    assert set(candidates) == {"a@x.example"}
    assert summary.candidates_ingested == 1


def test_deterministic_same_workbook_same_result(tmp_path: Path) -> None:
    sheets: dict[str, tuple[list[str], list[list[object]]]] = {
        "Beach": (
            BEACH_HEADERS,
            [
                [1, "A", "a@x.example", "Lead", "Java", "Bengaluru", "Yes", 5, "n"],
                [2, "B", "b@x.example", "Lead", "Kotlin", "Chennai", "No", 3, "n"],
            ],
        ),
        "Rolling Off": (
            ROLLING_HEADERS,
            [
                [
                    1,
                    "C",
                    "c@x.example",
                    "Lead",
                    "Java",
                    "Cl",
                    "2026-07-01",
                    "high",
                    "Chennai",
                    "No",
                    "n",
                ]
            ],
        ),
        "New Joiners": (
            JOINER_HEADERS,
            [[1, "D", "d@x.example", "Senior", "Python", "2026-07-11", "Bengaluru", "No", "n"]],
        ),
    }
    path = _write_workbook(tmp_path, sheets)

    first = ingest_candidates(path)
    second = ingest_candidates(path)
    assert first[0] == second[0]
    assert first[1] == second[1]
    assert list(first[0]) == list(second[0])  # key order stable too


# --- smoke test against the real shipped workbook --------------------------
REAL_WORKBOOK = Path(__file__).resolve().parents[2] / "data" / "demand-supply.xlsx"


@pytest.mark.skipif(not REAL_WORKBOOK.exists(), reason="real workbook not present")
def test_real_workbook_smoke() -> None:
    candidates, summary = ingest_candidates(REAL_WORKBOOK)

    # 10 Beach + 10 Rolling Off + 15 New Joiners = 35 (asserted so data drift fails).
    assert summary.candidates_ingested == 35
    assert len(candidates) == 35
    assert summary.issues == []
    assert summary.duplicate_emails_skipped == 0
    assert summary.blank_rows_skipped == 0

    by_source = {src: 0 for src in CandidateSource}
    for cand in candidates.values():
        by_source[cand.source] += 1
    assert by_source[CandidateSource.BEACH] == 10
    assert by_source[CandidateSource.ROLLING_OFF] == 10
    assert by_source[CandidateSource.NEW_JOINER] == 15
