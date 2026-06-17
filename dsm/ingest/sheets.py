"""Real candidate ingestion from the supply sheets (a-001-ingest-sheets).

Reads the three supply tabs (Beach / Rolling Off / New Joiners) of
``data/demand-supply.xlsx`` with openpyxl and maps each valid row to a frozen
``Candidate`` (``dsm/models.py``). Sheets-only: profile/feedback enrichment and
Open Roles mapping are out of scope for this feature.

The module is pure except for the single file read at the edge
(``ingest_candidates``); the parsing helpers below take plain values and never
touch the filesystem, network, clock or RNG (I-DET-1, ``docs/tech.md``).
"""

from __future__ import annotations

import re
from collections.abc import Iterable, Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Literal, NamedTuple, cast

from openpyxl import load_workbook
from pydantic import ValidationError

from dsm.ingest.models import IngestError, IngestSummary, RowIssue
from dsm.models import (
    AvailabilityState,
    Candidate,
    CandidateSource,
    FeedbackSignals,
    FreeNow,
    Location,
    NewJoiner,
    ProficiencyLevel,
    RollingOff,
    Skill,
)

_REMOTE_RE = re.compile(r"remote", re.IGNORECASE)
_CONFIDENCE_VALUES = ("high", "medium", "low")

# Header row is row 2; data begins at row 3 (I-LOAD-3).
_HEADER_ROW = 2
_FIRST_DATA_ROW = 3


class _SupplyTab(NamedTuple):
    """A supply tab to ingest, in fixed processing order (I-DET-1)."""

    sheet: str
    source: CandidateSource
    required: tuple[str, ...]  # headers without which the tab can't be mapped (I-LOAD-2)


# Fixed order: Beach → Rolling Off → New Joiners (I-EDGE-1 first-occurrence, I-DET-1).
_SUPPLY_TABS: tuple[_SupplyTab, ...] = (
    _SupplyTab("Beach", CandidateSource.BEACH, ("Name", "Email")),
    _SupplyTab(
        "Rolling Off",
        CandidateSource.ROLLING_OFF,
        ("Name", "Email", "Roll-off Date", "Confidence"),
    ),
    _SupplyTab("New Joiners", CandidateSource.NEW_JOINER, ("Name", "Email", "Join Date")),
)


def _norm(value: object) -> str:
    """Normalise a header/cell label for case-insensitive matching."""
    return str(value).strip().lower()


def parse_date(cell: object) -> date:
    """Coerce a date cell to a ``date`` (I-EDGE-3).

    Accepts a ``date``, a ``datetime`` (openpyxl's native type for date cells), or
    an ISO ``YYYY-MM-DD`` string. Anything else raises ``ValueError`` (caught by the
    reader and recorded as a ``RowIssue``).
    """
    # datetime is a subclass of date — check it first.
    if isinstance(cell, datetime):
        return cell.date()
    if isinstance(cell, date):
        return cell
    if isinstance(cell, str):
        text = cell.strip()
        try:
            return datetime.strptime(text, "%Y-%m-%d").date()
        except ValueError as exc:
            raise ValueError(f"unparseable date: {cell!r}") from exc
    raise ValueError(f"unparseable date: {cell!r}")


def parse_skills(raw: object) -> list[Skill]:
    """Split a *Key Skills* cell into normalised ``Skill`` records (I-CAND-3).

    Comma-separated, trimmed, lowercased, de-duplicated preserving first-seen
    order. Proficiency is not present in the sheets, so every skill defaults to
    ``INTERMEDIATE`` (OQ-2), to be overwritten when profiles supply real levels.
    An empty/``None`` cell yields ``[]``.
    """
    if raw is None:
        return []
    text = str(raw).strip()
    if not text:
        return []
    skills: list[Skill] = []
    seen: set[str] = set()
    for part in text.split(","):
        name = part.strip().lower()
        if not name or name in seen:
            continue
        seen.add(name)
        skills.append(Skill(name=name, proficiency=ProficiencyLevel.INTERMEDIATE))
    return skills


def parse_location(location_text: object, chennai_open: object) -> Location:
    """Map *Location* + *Chennai-open* to a ``Location`` (I-CAND-4, AD-020/063a).

    ``remote_eligible`` is ``True`` when *Chennai-open* is ``Yes`` **or** any
    location segment denotes remote (e.g. ``remote-India``, ``Remote (India)``).
    City is the first non-remote segment, or the first segment verbatim when every
    segment is remote. ``country`` defaults to ``"India"``.
    """
    text = "" if location_text is None else str(location_text).strip()
    segments = [seg.strip() for seg in text.split("/") if seg.strip()]

    open_to_chennai = chennai_open is not None and _norm(chennai_open) == "yes"
    has_remote_segment = any(_REMOTE_RE.search(seg) for seg in segments)
    remote_eligible = open_to_chennai or has_remote_segment

    city = ""
    if segments:
        non_remote = [seg for seg in segments if not _REMOTE_RE.search(seg)]
        city = non_remote[0] if non_remote else segments[0]

    return Location(city=city, country="India", remote_eligible=remote_eligible)


def _header_index(ws: object, *, sheet: str, required: Iterable[str]) -> dict[str, int]:
    """Map normalised header name → 1-based column index from the header row (row 2).

    Columns are resolved by name, not fixed position (I-LOAD-3). Raises
    ``IngestError`` naming the offending column if a required header is missing
    (I-LOAD-2).
    """
    headers: dict[str, int] = {}
    for col, cell in enumerate(ws[_HEADER_ROW], start=1):  # type: ignore[index]
        if cell.value is None:
            continue
        headers[_norm(cell.value)] = col

    missing = [name for name in required if _norm(name) not in headers]
    if missing:
        raise IngestError(f"{sheet}: missing required column(s): {', '.join(missing)}")
    return headers


def _is_blank(values: Iterable[object]) -> bool:
    """True when every cell in a data row is empty/``None`` (I-EDGE-2)."""
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def _cell(values: Sequence[object], headers: dict[str, int], name: str) -> object:
    """Look up a cell value by header name (1-based index), or ``None`` if absent."""
    idx = headers.get(_norm(name))
    if idx is None or idx > len(values):
        return None
    return values[idx - 1]


def _require_text(values: Sequence[object], headers: dict[str, int], name: str, label: str) -> str:
    """Return a required non-empty string cell, else raise ``ValueError`` (I-VAL-1)."""
    raw = _cell(values, headers, name)
    text = "" if raw is None else str(raw).strip()
    if not text:
        raise ValueError(f"missing {label}")
    return text


def _availability(
    values: Sequence[object], headers: dict[str, int], source: CandidateSource
) -> AvailabilityState:
    """Build the availability variant from sheet membership (I-CAND-2/5).

    Beach → ``FreeNow``; Rolling Off → ``RollingOff(expected_date, confidence)``;
    New Joiners → ``NewJoiner(join_date)``. Bad/missing dates raise ``ValueError``
    and a confidence outside ``{high, medium, low}`` raises ``ValueError`` — both
    caught by the reader and recorded as a ``RowIssue``.
    """
    if source is CandidateSource.BEACH:
        return FreeNow()
    if source is CandidateSource.ROLLING_OFF:
        expected = parse_date(_cell(values, headers, "Roll-off Date"))
        raw_conf = _cell(values, headers, "Confidence")
        conf = "" if raw_conf is None else str(raw_conf).strip().lower()
        if conf not in _CONFIDENCE_VALUES:
            raise ValueError(f"bad confidence: {raw_conf!r}")
        return RollingOff(
            expected_date=expected,
            confidence=cast(Literal["high", "medium", "low"], conf),
        )
    # New joiner (CandidateSource.NEW_JOINER)
    return NewJoiner(join_date=parse_date(_cell(values, headers, "Join Date")))


def _row_to_candidate(
    values: Sequence[object], headers: dict[str, int], source: CandidateSource
) -> Candidate:
    """Map one valid supply row to a ``Candidate`` (I-CAND-1/6).

    Raises ``ValueError`` (missing email/name, bad date/confidence) or pydantic
    ``ValidationError``; both are caught by the reader and turned into a
    ``RowIssue``. ``feedback`` is left empty and enrichment fields stay ``None``
    (out of scope here). New-joiner uncertainty is represented by ``source`` (OQ-1).
    """
    email = _require_text(values, headers, "Email", "email")
    name = _require_text(values, headers, "Name", "name")
    location = parse_location(
        _cell(values, headers, "Location"),
        _cell(values, headers, "Chennai-open"),
    )
    raw_skills = _cell(values, headers, "Key Skills")
    if raw_skills is None:
        raw_skills = _cell(values, headers, "Key Skills (from CV)")
    return Candidate(
        email=email,
        name=name,
        location=location,
        availability=_availability(values, headers, source),
        skills=parse_skills(raw_skills),
        feedback=FeedbackSignals(),
        source=source,
    )


def _read_email(values: Sequence[object], headers: dict[str, int]) -> str | None:
    """Best-effort email read for a ``RowIssue`` on an otherwise-invalid row."""
    raw = _cell(values, headers, "Email")
    text = "" if raw is None else str(raw).strip()
    return text or None


def ingest_candidates(
    path: str | Path,
) -> tuple[dict[str, Candidate], IngestSummary]:
    """Load candidate supply sheets into typed ``Candidate`` records (I-LOAD-1).

    Opens ``path`` once with openpyxl (``data_only=True``) and iterates the three
    supply tabs in fixed order (Beach → Rolling Off → New Joiners); any other tab
    (e.g. ``Open Roles``) is ignored. Blank rows are skipped, duplicate emails keep
    the first occurrence, and any row that fails validation becomes a ``RowIssue``
    rather than aborting the run (I-VAL-1, I-EDGE-1/2/3). A missing tab or required
    header is fatal (``IngestError``, I-LOAD-2).

    Returns ``(candidates_by_email, summary)`` — no ``IngestResult`` wrapper (OQ-5).
    Deterministic: same workbook → same result, no clock/RNG (I-DET-1).
    """
    workbook_path = str(path)
    wb = load_workbook(workbook_path, data_only=True)

    candidates: dict[str, Candidate] = {}
    first_seen: dict[str, tuple[str, int]] = {}  # email → (sheet, row) of first occurrence
    issues: list[RowIssue] = []
    rows_seen = 0
    blank_skipped = 0
    duplicate_skipped = 0

    for tab in _SUPPLY_TABS:
        if tab.sheet not in wb.sheetnames:
            raise IngestError(f"missing required supply tab: {tab.sheet!r}")
        ws = wb[tab.sheet]
        headers = _header_index(ws, sheet=tab.sheet, required=tab.required)

        for offset, values in enumerate(ws.iter_rows(min_row=_FIRST_DATA_ROW, values_only=True)):
            row_number = _FIRST_DATA_ROW + offset
            if _is_blank(values):
                blank_skipped += 1
                continue

            rows_seen += 1
            try:
                candidate = _row_to_candidate(values, headers, tab.source)
            except (ValueError, ValidationError) as exc:
                issues.append(
                    RowIssue(
                        sheet=tab.sheet,
                        row_number=row_number,
                        email=_read_email(values, headers),
                        reason=str(exc).replace("\n", " "),
                    )
                )
                continue

            prior = first_seen.get(candidate.email)
            if prior is not None:
                duplicate_skipped += 1
                issues.append(
                    RowIssue(
                        sheet=tab.sheet,
                        row_number=row_number,
                        email=candidate.email,
                        reason=(
                            "duplicate email; kept first occurrence from "
                            f"{prior[0]} row {prior[1]}"
                        ),
                    )
                )
                continue

            first_seen[candidate.email] = (tab.sheet, row_number)
            candidates[candidate.email] = candidate

    summary = IngestSummary(
        workbook_path=workbook_path,
        candidate_rows_seen=rows_seen,
        candidates_ingested=len(candidates),
        blank_rows_skipped=blank_skipped,
        duplicate_emails_skipped=duplicate_skipped,
        issues=issues,
    )
    return candidates, summary
